"""M38.4 — Excel parser.

Each sheet becomes one ``DocumentChunk``. The sheet's header row is kept
verbatim (so the chunk reads like a real CSV head), and the next
``PREVIEW_ROWS`` body rows are joined with newlines. Cells that contain
formulas keep their ``=EXPR`` text (``data_only=False``) — this is
intentional, formulas are *better* for search than computed values
(spec §"开放问题 4"). Empty sheets are skipped (no zero-length chunks).

The parser is **authoritative** for chunking: it returns its own
``chunks`` list rather than letting ``DocumentParser._create_chunks``
split the joined text a second time. The class attribute
``preserves_chunks = True`` tells ``DocumentParser.parse()`` to skip the
secondary split. Each chunk dict carries the fields
``DocumentChunk`` needs (modality / sheet_name / page_number /
image_caption) plus a ``chunk_metadata`` JSON body for the
``chunk_metadata`` column.

Image / chart / pivot objects inside a sheet are **not** extracted —
spec §2 explicitly defers Excel images to v2. The spec also caps body
preview at 50 rows; this avoids one mega-chunk for a 100k-row sheet.
"""
from __future__ import annotations

import os
from typing import Any, Dict, List

from . import BaseParser


# Spec §"开放问题 4" / §1.3: pre-M38.4 Excel uploads turned the whole
# sheet into one giant chunk. The spec caps the preview at 50 rows
# per sheet — bigger sheets are still searchable via the header + first
# 50 rows (document_tasks writes the row_count into chunk_metadata so
# the UI can surface "showing 50 of 1,234 rows" if needed).
PREVIEW_ROWS = 50


class ExcelParser(BaseParser):
    """Excel (.xlsx / .xls) parser — one chunk per sheet."""

    #: Tell ``DocumentParser`` this parser already produced final chunks;
    #: the secondary text-split pass should be skipped.
    preserves_chunks: bool = True

    def get_type(self) -> str:
        return "excel"

    def parse(self, file_path: str) -> Dict[str, Any]:
        try:
            # openpyxl ships no type stubs (no ``types-openpyxl``) — flag
            # the import so mypy doesn't complain at parse time. Same
            # pattern as the docling / pymilvus imports elsewhere in
            # the codebase.
            import openpyxl  # type: ignore[import-untyped]
            # local import — heavy dep; not needed for text/pdf
        except ImportError as exc:  # pragma: no cover - dependency miss
            return self._fallback_parse(file_path, reason=f"openpyxl not installed: {exc}")

        try:
            # ``data_only=False`` keeps the formula string ``=SUM(A1:A10)``
            # rather than the cached computed value. Spec §"开放问题 4"
            # calls this out: formulas carry more retrieval signal than
            # numbers (a user searching "sum of price" wants to find the
            # formula text, not the result).
            workbook = openpyxl.load_workbook(file_path, data_only=False)
        except Exception as exc:
            return self._fallback_parse(file_path, reason=f"openpyxl: {type(exc).__name__}: {exc}")

        chunk_texts: List[str] = []
        chunk_records: List[Dict[str, Any]] = []
        sheet_count = 0

        for worksheet in workbook.worksheets:
            sheet_count += 1
            sheet_name = str(worksheet.title or "")
            # ``iter_rows(values_only=True)`` materialises every cell;
            # for sheets with hundreds of thousands of rows this would
            # be a memory spike. Spec caps at 50 rows preview, so we
            # materialise once and slice. The full sheet is *not* read
            # into memory beyond what openpyxl already loads (one
            # ``Worksheet`` object per sheet — typically OK for KB-sized
            # Excels < 50 MB).
            all_rows = list(worksheet.iter_rows(values_only=True))
            if not all_rows:
                # Empty sheet — skip; emitting a zero-text chunk would
                # poison the vector store with a no-op entry.
                continue

            # Header: first non-empty row, defensively. ``all_rows[0]``
            # is the canonical Excel header position; if it's a fully
            # blank row we still keep going (spec says header is row 0,
            # not "first non-empty row" — users may want an explicit
            # empty header).
            header_row = all_rows[0]
            header_cells = [_cell_to_text(c) for c in header_row]
            body_rows = all_rows[1:]

            # Cap body preview; the full row_count goes to metadata so
            # callers know they only got a sample.
            preview = body_rows[:PREVIEW_ROWS]
            preview_lines = [",".join(header_cells)]
            for row in preview:
                preview_lines.append(",".join(_cell_to_text(c) for c in row))

            # Header line plus sheet name as the chunk title — search
            # for "Sheet: 报价单" should match even if all body rows are
            # empty (header alone is enough context for the KB).
            chunk_text = f"Sheet: {sheet_name}\n" + "\n".join(preview_lines)
            chunk_texts.append(chunk_text)

            chunk_records.append(
                {
                    "content": chunk_text,
                    "chunk_index": len(chunk_records),
                    "length": len(chunk_text),
                    "strategy": "excel_sheet",
                    "modality": "text",
                    "sheet_name": sheet_name,
                    "page_number": None,
                    "image_caption": None,
                    "chunk_metadata": {
                        "kind": "excel_sheet",
                        "sheet_name": sheet_name,
                        "row_count": len(body_rows),
                        "header": header_cells,
                        "preview_rows": len(preview),
                    },
                }
            )

        # Concatenated text — preserved for legacy callers that only
        # look at ``result["text"]`` (e.g. the manual fallback in
        # ``document_tasks``). Double newline separates sheets so
        # accidental text-only reads still keep boundaries.
        joined_text = "\n\n".join(chunk_texts)
        return {
            "text": joined_text,
            "metadata": {
                "type": self.get_type(),
                "format": "xlsx",
                "sheet_count": sheet_count,
                "file_path": file_path,
                "chunk_count": len(chunk_records),
            },
            "chunks": chunk_records,
            "chunk_metadata": [c["chunk_metadata"] for c in chunk_records],
        }


def _cell_to_text(cell: Any) -> str:
    """Coerce an openpyxl cell value to a search-friendly string.

    ``None`` → empty string (NOT the literal "None" — openpyxl hands
    back ``None`` for blank cells, and Python's default ``str(None)``
    would poison chunks with the word "None" everywhere).
    Numbers, booleans, ``datetime`` objects all stringify sensibly
    via ``str()``. ``datetime`` cells are common in Excel financial
    sheets — let them stringify to ISO 8601.
    """
    if cell is None:
        return ""
    if isinstance(cell, str):
        return cell
    return str(cell)


__all__ = ["ExcelParser"]
