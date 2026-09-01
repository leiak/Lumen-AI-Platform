"""M38.4 Step 4 — ExcelParser unit tests.

Real xlsx fixtures via openpyxl (the parser's own dep). Tests cover:

- multi-sheet produces one chunk per non-empty sheet
- empty sheets are skipped (no zero-length chunks)
- formula cells keep their ``=EXPR`` text (``data_only=False``)
- Chinese / unicode headers round-trip cleanly
- 50-row preview cap (sheet with 200 rows yields exactly 50 body lines)
- ``preserves_chunks=True`` flag is set so ``DocumentParser._create_chunks``
  doesn't shred sheet boundaries with the secondary text-split
- ``get_type`` returns ``"excel"`` for the factory's ``PARSERS`` lookup

We never mock openpyxl — the parser's contract IS to drive openpyxl,
so a real fixture is the most honest test. The fixture cost is small
(<< 100 ms per test).
"""
from __future__ import annotations

from typing import List

import pytest
from openpyxl import Workbook

from lumen_services.parsers import ExcelParser
from lumen_services.parsers.excel_parser import PREVIEW_ROWS


# ----------------------------------------------------------------------
# Fixture builders
# ----------------------------------------------------------------------


def _write_xlsx(path: str, sheets: List[dict]) -> None:
    """Write a small xlsx with one or more sheets.

    ``sheets`` is a list of ``{"name": str, "rows": List[List[cell]]}``;
    the first row of each sheet is treated as the header.
    """
    workbook = Workbook()
    # Remove the auto-created "Sheet" — we'll create sheets from the spec.
    workbook.remove(workbook.active)
    for spec in sheets:
        worksheet = workbook.create_sheet(title=spec["name"])
        for row in spec["rows"]:
            worksheet.append(row)
    workbook.save(path)


def _make_simple_workbook(path: str) -> None:
    _write_xlsx(
        path,
        [
            {"name": "Sales", "rows": [
                ["Item", "Price", "Qty"],
                ["Apple", 1.50, 10],
                ["Banana", 0.80, 20],
            ]},
            {"name": "Inventory", "rows": [
                ["SKU", "Stock"],
                ["A-001", 100],
                ["B-002", 50],
            ]},
        ],
    )


# ----------------------------------------------------------------------
# Tests
# ----------------------------------------------------------------------


def test_excel_parser_returns_one_chunk_per_sheet(tmp_path):
    out = tmp_path / "two_sheets.xlsx"
    _make_simple_workbook(str(out))
    result = ExcelParser().parse(str(out))

    assert result["metadata"]["type"] == "excel"
    assert result["metadata"]["sheet_count"] == 2
    assert len(result["chunks"]) == 2

    # chunk 0 → Sales, chunk 1 → Inventory
    assert result["chunks"][0]["sheet_name"] == "Sales"
    assert result["chunks"][1]["sheet_name"] == "Inventory"

    # modality is always 'text' for Excel — images inside sheets are
    # deferred to v2 (spec §2).
    for c in result["chunks"]:
        assert c["modality"] == "text"
        assert c["page_number"] is None
        assert c["image_caption"] is None


def test_excel_parser_chunk_text_contains_header_and_rows(tmp_path):
    out = tmp_path / "sales.xlsx"
    _make_simple_workbook(str(out))
    result = ExcelParser().parse(str(out))

    sales_chunk_text = result["chunks"][0]["content"]
    assert "Sheet: Sales" in sales_chunk_text
    # CSV header line
    assert "Item,Price,Qty" in sales_chunk_text
    # Body rows
    assert "Apple,1.5,10" in sales_chunk_text
    assert "Banana,0.8,20" in sales_chunk_text


def test_excel_parser_skips_empty_sheets(tmp_path):
    out = tmp_path / "with_empty.xlsx"
    _write_xlsx(
        str(out),
        [
            {"name": "Real", "rows": [["a", "b"], ["1", "2"]]},
            {"name": "Empty", "rows": [[]]},  # 0 body rows
            {"name": "Another", "rows": [["x"], ["y"]]},
        ],
    )
    result = ExcelParser().parse(str(out))

    # Empty sheet is skipped — we get 2 chunks, not 3.
    assert result["metadata"]["sheet_count"] == 3  # openpyxl still counts it
    assert len(result["chunks"]) == 2
    sheet_names = [c["sheet_name"] for c in result["chunks"]]
    assert "Empty" not in sheet_names


def test_excel_parser_preserves_formula_strings(tmp_path):
    """Spec §"开放问题 4" — formulas carry more retrieval signal than
    computed values, so the parser must keep ``=EXPR`` text rather
    than the cached numeric result."""
    out = tmp_path / "formulas.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    ws = workbook.create_sheet("Math")
    ws.append(["A", "B", "Sum"])
    ws.append([1, 2, "=A2+B2"])
    ws.append([3, 4, "=A3+B3"])
    workbook.save(str(out))

    result = ExcelParser().parse(str(out))
    math_text = result["chunks"][0]["content"]
    assert "=A2+B2" in math_text
    assert "=A3+B3" in math_text
    # The computed result (3, 7) must NOT replace the formula string.
    # If ``data_only=True`` were used by accident these would appear.
    assert ",3," not in math_text
    assert ",7," not in math_text


def test_excel_parser_chinese_headers(tmp_path):
    """Unicode round-trip — Chinese column headers must survive."""
    out = tmp_path / "chinese.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    ws = workbook.create_sheet("产品")
    ws.append(["名称", "数量", "价格"])
    ws.append(["产品A", 10, 100])
    ws.append(["产品B", 20, 200])
    workbook.save(str(out))

    result = ExcelParser().parse(str(out))
    chunk_text = result["chunks"][0]["content"]
    assert "Sheet: 产品" in chunk_text
    assert "名称,数量,价格" in chunk_text
    assert result["chunks"][0]["sheet_name"] == "产品"


def test_excel_parser_caps_preview_at_50_rows(tmp_path):
    """Spec §"开放问题 4": large sheets only contribute 50 body rows."""
    big_sheet_rows = [["idx", "val"]] + [[i, i * 2] for i in range(200)]
    out = tmp_path / "big.xlsx"
    _write_xlsx(
        str(out),
        [{"name": "BigSheet", "rows": big_sheet_rows}],
    )
    result = ExcelParser().parse(str(out))

    chunk = result["chunks"][0]
    # 50 body lines + 1 header line + "Sheet: BigSheet\n" line
    assert chunk["chunk_metadata"]["preview_rows"] == PREVIEW_ROWS == 50
    assert chunk["chunk_metadata"]["row_count"] == 200  # total body, not preview
    # The content must contain idx=49 (last preview row, "49,98") but
    # not idx=50 (which would render as "50,100" — beyond the cap).
    assert "49,98" in chunk["content"]
    assert "50,100" not in chunk["content"]


def test_excel_parser_preserves_chunks_flag_is_true():
    """``DocumentParser.parse()`` honours ``preserves_chunks`` to skip
    the secondary text-split for parsers that already produced final
    chunks. The flag must be ``True`` for the multimodal parsers so
    sheet boundaries survive."""
    assert ExcelParser.preserves_chunks is True


def test_excel_parser_get_type():
    """``get_type`` must return ``"excel"`` for factory dispatch."""
    assert ExcelParser().get_type() == "excel"


def test_excel_parser_metadata_records_chunk_count(tmp_path):
    out = tmp_path / "two.xlsx"
    _make_simple_workbook(str(out))
    result = ExcelParser().parse(str(out))
    # metadata.chunk_count should mirror len(chunks) for downstream
    # observability (logs / notifications).
    assert result["metadata"]["chunk_count"] == 2


def test_excel_parser_handles_missing_file(tmp_path):
    """A missing xlsx must fall back to the legacy raw-text path rather
    than raising. The result keeps ``metadata.parse_error`` so the
    caller can mark the Document as failed."""
    result = ExcelParser().parse(str(tmp_path / "missing.xlsx"))
    assert result["metadata"]["fallback"] is True
    assert "parse_error" in result["metadata"]
    assert "openpyxl" in result["metadata"]["parse_error"]


def test_excel_parser_returns_chunk_metadata_list_aligned_with_chunks(tmp_path):
    """``chunk_metadata`` list length must match ``chunks`` list length
    and the i-th entry must describe the i-th chunk."""
    out = tmp_path / "aligned.xlsx"
    _make_simple_workbook(str(out))
    result = ExcelParser().parse(str(out))

    chunks = result["chunks"]
    chunk_meta = result["chunk_metadata"]
    assert len(chunks) == len(chunk_meta) == 2

    for chunk, meta in zip(chunks, chunk_meta):
        # meta carries the same sheet_name as the chunk.
        assert meta["sheet_name"] == chunk["sheet_name"]
        assert meta["kind"] == "excel_sheet"
        # header is a list of cell texts matching the column count.
        assert isinstance(meta["header"], list)
        assert len(meta["header"]) >= 1
